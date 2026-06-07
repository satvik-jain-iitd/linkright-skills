"""
Update Job Database.xlsx with full pipeline results.

Columns 1-4: Company / Product / Locations / Why (from source data)
Columns 5-6: Career URL + Portal Type (from career_urls_result.json)
Columns 7-8: Validation Status + Jobs Found (from jobs.db)

Handles both:
  - Existing 300 companies in the Excel — refresh cols 5-8
  - New 298 from Next 300.md — APPEND to appropriate sheet with full metadata
"""

import json
import re
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

BASE = Path(__file__).parent.parent
EXCEL_PATH = BASE / "Job Database.xlsx"
JSON_PATH = BASE / "db" / "career_urls_result.json"
DB_PATH = BASE / "db" / "jobs.db"
MD_PATH = BASE / "Next 300.md"

# Segment → Excel sheet
SEGMENT_SHEET = {
    "india_mnc": "Big Tech & MNCs - India",
    "global_ai_startup": "Consumer AI Startups - Global",
    "dubai_visa": "Dubai - Visa Sponsors",
}

SHEET_FILLS = {
    "Big Tech & MNCs - India":        PatternFill("solid", fgColor="1F4E79"),
    "Consumer AI Startups - Global":  PatternFill("solid", fgColor="1E6B3C"),
    "Dubai - Visa Sponsors":          PatternFill("solid", fgColor="7B3F00"),
}

# Name mapping: Excel sheet name → JSON key (for first 300)
NAME_MAP = {
    "hp inc. (hewlett-packard)": "hp inc.",
    "deloitte usi (tech arm)": "deloitte usi",
    "citibank (citi tech hubs)": "citi (citibank tech)",
    "tata digital (tata neu)": "tata neu (tata digital)",
    "info edge (naukri/jeevansathi)": "info edge (naukri)",
    "dream11 (sporta tech)": "dream11",
    "limitless (rewind)": "limitless ai",
    "decart (oasis)": "decart ai (oasis)",
    "lex": "lex (every)",
    "meteor": "meteor ai",
    "cocreate": "cocreate ai",
    "interface": "interface ai",
    "clay": "clay (crm)",
    "genspark (mainconcept)": "genspark",
    "codiumai": "codiumai (qodo)",
    "linum": "linum ai",
    "voicenotes.com": "voicenotes",
    "mindy": "mindy ai",
    "hedra": "hedra ai",
    "kling ai": "kling ai (kuaishou)",
    "lmarena (chatbot arena)": "lm arena (lmsys)",
    "dubizzle (empg group)": "dubizzle (empg)",
    "deliveroo": "deliveroo uae",
    "mashreq neo": "mashreq neo (mashreq)",
    "al tayer digital": "al tayer digital (ounass)",
    "qlub -": "qlub",
    "liv. bank": "liv. bank (emirates nbd)",
    "life pharmacy llc": "life pharmacy",
    "seva stories (seva)": "seva stories",
    "mamo (mamo pay)": "mamo pay",
}


def portal_fill(ptype):
    if not ptype: return None
    if ptype.startswith("ATS_Workday"):    return PatternFill("solid", fgColor="E2EFDA")
    if ptype.startswith("ATS_Greenhouse"): return PatternFill("solid", fgColor="E2EFDA")
    if ptype.startswith("ATS_Lever"):      return PatternFill("solid", fgColor="E2EFDA")
    if ptype.startswith("ATS_Ashby"):      return PatternFill("solid", fgColor="E2EFDA")
    if ptype.startswith("ATS_"):           return PatternFill("solid", fgColor="FFF2CC")
    if ptype == "Tier2_CSR_SPA":           return PatternFill("solid", fgColor="DEEBF7")
    if ptype == "Tier1_StaticHTML":        return PatternFill("solid", fgColor="F2F2F2")
    if ptype == "Tier3_ProtectedAPI":      return PatternFill("solid", fgColor="FCE4D6")
    return None


def status_fill(status):
    if not status: return None
    if status == "PASS":            return PatternFill("solid", fgColor="C6EFCE")  # green
    if "GEO_BLOCKED" in status:     return PatternFill("solid", fgColor="FFEB9C")  # yellow
    if "ANTI_BOT" in status:        return PatternFill("solid", fgColor="FFEB9C")
    if status.startswith("FAIL"):   return PatternFill("solid", fgColor="FFC7CE")  # red
    return None


def parse_next300(path: Path) -> dict[str, dict]:
    """Returns: company_name → {segment, product, locations, why}"""
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8")
    segment_keys = {1: "india_mnc", 2: "global_ai_startup", 3: "dubai_visa"}
    rows = {}
    current_segment = None
    for line in text.split("\n"):
        m_seg = re.match(r"##\s+\*\*Segment\s+(\d+)", line)
        if m_seg:
            current_segment = int(m_seg.group(1))
            continue
        if current_segment and line.strip().startswith("|") and "**" in line:
            cells = [c.strip() for c in line.split("|")[1:-1]]
            if len(cells) < 5: continue
            if cells[0].lower().startswith("company name"): continue
            m_name = re.search(r"\*\*([^*]+)\*\*", cells[0])
            if not m_name: continue
            company = m_name.group(1).strip()
            rows[company] = {
                "segment": segment_keys.get(current_segment, "unknown"),
                "product": cells[1] if len(cells) > 1 else "",
                "locations": cells[2] if len(cells) > 2 else "",
                "why": cells[3] if len(cells) > 3 else "",
            }
    return rows


def main():
    # Load all data sources
    with open(JSON_PATH) as f:
        entries = json.load(f)
    lookup_url = {e["company"].lower().strip(): e for e in entries}

    conn = sqlite3.connect(DB_PATH)
    lookup_db = {}
    for r in conn.execute(
        "SELECT company, validation_status, jobs_found, verified, portal_type, career_url FROM companies"
    ):
        lookup_db[r[0].lower().strip()] = {
            "validation_status": r[1],
            "jobs_found": r[2] or 0,
            "verified": r[3] or 0,
            "portal_type": r[4],
            "career_url": r[5],
        }
    conn.close()

    next300_meta = parse_next300(MD_PATH)
    # Track which Next 300 companies already exist in Excel (to skip)
    existing_excel_companies = set()

    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- Step 1: Update existing rows in each sheet (cols 5-8) ---
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Reset all columns 5-8 (re-run safety)
        while ws.max_column > 4:
            ws.delete_cols(5)

        hdr_fill = SHEET_FILLS.get(sheet_name)
        hdr_font = Font(bold=True, color="FFFFFF")

        for col_idx, hdr in [(5, "Career Portal URL"), (6, "Portal Type"),
                              (7, "Validation Status"), (8, "Jobs Found")]:
            cell = ws.cell(1, col_idx)
            cell.value = hdr
            cell.font = hdr_font
            if hdr_fill:
                cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        matched = 0
        for row in range(2, ws.max_row + 1):
            company_cell = ws.cell(row, 1).value
            if not company_cell:
                continue
            existing_excel_companies.add(company_cell.lower().strip())

            key = company_cell.lower().strip()
            entry = lookup_url.get(key) or lookup_url.get(NAME_MAP.get(key, ""))
            db = lookup_db.get(key) or lookup_db.get(NAME_MAP.get(key, ""))

            career_url = (entry or {}).get("career_url", "") if entry else (db or {}).get("career_url", "")
            portal_type = (entry or {}).get("portal_type", "Not Found") if entry else (db or {}).get("portal_type", "Not Found")
            v_status = (db or {}).get("validation_status", "")
            jobs = (db or {}).get("jobs_found", 0)

            ws.cell(row, 5).value = career_url
            ws.cell(row, 5).alignment = Alignment(wrap_text=True, vertical="center")

            tc = ws.cell(row, 6)
            tc.value = portal_type
            tc.alignment = Alignment(vertical="center")
            f = portal_fill(portal_type)
            if f: tc.fill = f

            sc = ws.cell(row, 7)
            sc.value = v_status or ""
            sc.alignment = Alignment(vertical="center", horizontal="center")
            f = status_fill(v_status)
            if f: sc.fill = f

            jc = ws.cell(row, 8)
            jc.value = jobs
            jc.alignment = Alignment(vertical="center", horizontal="center")

            if career_url:
                matched += 1
        print(f"  [{sheet_name}] refreshed {matched}/{ws.max_row-1} existing rows")

    # --- Step 2: Append new 298 from Next 300.md ---
    appended_by_sheet = {sn: 0 for sn in SEGMENT_SHEET.values()}
    for company, meta in next300_meta.items():
        key = company.lower().strip()
        if key in existing_excel_companies:
            continue  # already in sheet (e.g. Epicenter dup)

        sheet_name = SEGMENT_SHEET.get(meta["segment"])
        if not sheet_name:
            continue
        ws = wb[sheet_name]

        entry = lookup_url.get(key) or {}
        db = lookup_db.get(key) or {}

        new_row = ws.max_row + 1
        ws.cell(new_row, 1).value = company
        ws.cell(new_row, 2).value = meta["product"]
        ws.cell(new_row, 3).value = meta["locations"]
        ws.cell(new_row, 4).value = meta["why"]
        ws.cell(new_row, 5).value = entry.get("career_url", db.get("career_url", ""))

        ptype = entry.get("portal_type", db.get("portal_type", "Not Found"))
        tc = ws.cell(new_row, 6)
        tc.value = ptype
        f = portal_fill(ptype)
        if f: tc.fill = f

        v = db.get("validation_status", "")
        sc = ws.cell(new_row, 7)
        sc.value = v
        f = status_fill(v)
        if f: sc.fill = f

        ws.cell(new_row, 8).value = db.get("jobs_found", 0)

        for c in range(1, 9):
            ws.cell(new_row, c).alignment = Alignment(
                wrap_text=(c in (2, 4)),
                vertical="center",
                horizontal="center" if c in (7, 8) else "left",
            )

        appended_by_sheet[sheet_name] += 1

    for sn, n in appended_by_sheet.items():
        print(f"  [{sn}] appended {n} new rows from Next 300.md")

    # Column widths
    for sn in wb.sheetnames:
        ws = wb[sn]
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 11

    wb.save(EXCEL_PATH)
    print(f"\n✓ Saved → {EXCEL_PATH}")

    # Final tally per sheet
    print("\n=== Final Excel state ===")
    wb2 = openpyxl.load_workbook(EXCEL_PATH)
    for sn in wb2.sheetnames:
        print(f"  {sn}: {wb2[sn].max_row - 1} companies")


if __name__ == "__main__":
    main()
